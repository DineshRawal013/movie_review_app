"""
RTM.md → RTM.xlsx converter
Produces a fully-formatted Excel workbook with:
  - One sheet per requirement module + a Summary sheet
  - Frozen header row, auto-filter dropdowns
  - Color-coded Status column
  - Auto-fit column widths
  - Charts on Summary sheet for pass/fail rates
"""

import re
import sys
import io
from pathlib import Path

# Force UTF-8 stdout on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ── Paths ───────────────────────────────────────────────────────────────────
DOCS = Path(__file__).parent.parent
RTM_PATH = DOCS / "RTM.md"
OUT_PATH = Path(__file__).parent / "RTM.xlsx"

# ── Colour palette ───────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
PASS_FILL     = PatternFill("solid", fgColor="C6EFCE")   # light green
FAIL_FILL     = PatternFill("solid", fgColor="FFC7CE")   # light red
INDEV_FILL    = PatternFill("solid", fgColor="FFEB9C")   # light yellow
DEFERRED_FILL = PatternFill("solid", fgColor="D9D9D9")   # light grey
PARTIAL_FILL  = PatternFill("solid", fgColor="FFEB9C")   # same as in-dev
ALT_FILL      = PatternFill("solid", fgColor="F2F7FF")   # very light blue (alternating rows)

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT   = Font(name="Calibri", size=10)
TITLE_FONT  = Font(name="Calibri", bold=True, size=14, color="1F3864")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Status → fill mapping ────────────────────────────────────────────────────
def status_fill(status: str) -> PatternFill:
    s = status.upper()
    if "PASS" in s:           return PASS_FILL
    if "FAIL" in s:           return FAIL_FILL
    if "DEFERRED" in s:       return DEFERRED_FILL
    if "IN-DEV" in s or "IN DEV" in s: return INDEV_FILL
    if "PARTIAL" in s:        return PARTIAL_FILL
    if "PLANNED" in s:        return INDEV_FILL
    return PatternFill()       # no fill

# ── Markdown table parser ────────────────────────────────────────────────────
def parse_md_tables(text: str):
    """
    Returns list of dicts:
      { 'section': str, 'headers': [str], 'rows': [[str]] }
    """
    tables = []
    current_section = "General"
    lines = text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        # track section heading
        m = re.match(r'^#{2,4}\s+(.+)', line)
        if m:
            current_section = m.group(1).strip()
            i += 1
            continue
        # detect table header row (contains | characters)
        if '|' in line and i + 1 < len(lines) and re.match(r'\s*\|[-| :]+\|\s*$', lines[i+1]):
            headers = [c.strip() for c in line.strip().strip('|').split('|')]
            i += 2  # skip separator
            rows = []
            while i < len(lines) and '|' in lines[i]:
                cells = [c.strip() for c in lines[i].strip().strip('|').split('|')]
                # pad to header length
                while len(cells) < len(headers):
                    cells.append('')
                rows.append(cells[:len(headers)])
                i += 1
            if rows:
                tables.append({
                    'section': current_section,
                    'headers': headers,
                    'rows': rows
                })
            continue
        i += 1
    return tables


def clean(text: str) -> str:
    """Strip markdown bold/italic/links for cell text."""
    text = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', text)   # [text](url)
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)             # **bold**
    text = re.sub(r'\*(.+?)\*', r'\1', text)                  # *italic*
    text = re.sub(r'`(.+?)`', r'\1', text)                    # `code`
    return text.strip()


# ── Sheet builder ────────────────────────────────────────────────────────────
def write_sheet(ws, table: dict, sheet_title: str):
    headers = table['headers']
    rows    = table['rows']

    # Title row
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=sheet_title)
    title_cell.font      = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Header row
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[2].height = 30

    # Data rows
    for row_idx, row_data in enumerate(rows, start=3):
        alt = (row_idx % 2 == 0)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=clean(value))
            cell.font      = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border    = BORDER
            # status column coloring
            if headers[col_idx - 1].lower() in ("status", "result", "test result"):
                fill = status_fill(value)
                if fill.patternType:
                    cell.fill = fill
            elif alt:
                cell.fill = ALT_FILL

    # Auto-filter on header row
    ws.auto_filter.ref = (
        f"A2:{get_column_letter(len(headers))}{ len(rows) + 2 }"
    )
    # Freeze header rows
    ws.freeze_panes = "A3"

    # Auto-fit column widths (capped at 60)
    for col_idx, h in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(h)
        for row_data in rows:
            if col_idx - 1 < len(row_data):
                max_len = max(max_len, min(len(clean(row_data[col_idx-1])), 60))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)


# ── Summary sheet with charts ────────────────────────────────────────────────
SECTION_MAP = {
    "Movie Browsing and Discovery": "Browsing",
    "Movie Search and Filtering":   "Search",
    "Movie Detail Page":            "Detail",
    "User Authentication":          "Auth",
    "Star Ratings":                 "Ratings",
    "Text Reviews":                 "Reviews",
    "User Profile":                 "Profile",
    "Admin — Review Moderation":    "Admin-Mod",
    "Admin — Movie Management":     "Admin-Movie",
    "Non-Functional Requirements Traceability": "NFR",
}


def build_summary(wb, all_tables):
    ws = wb.create_sheet("Summary", 0)

    # Header
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "Movie Review App — RTM v2.1 — Test Execution Summary"
    c.font      = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Sub-header
    sub_headers = ["Module", "Total Reqs", "PASS", "FAIL", "DEFERRED", "Pass Rate"]
    for col, h in enumerate(sub_headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill      = PatternFill("solid", fgColor="2E75B6")
        cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = BORDER

    # Compute per-module stats
    summary_rows = []
    for tbl in all_tables:
        sec = tbl['section']
        label = SECTION_MAP.get(sec, sec[:20])
        if 'Status' not in tbl['headers'] and 'Result' not in tbl['headers']:
            continue
        try:
            status_col = next(
                i for i, h in enumerate(tbl['headers'])
                if h.lower() in ('status', 'result', 'test result')
            )
        except StopIteration:
            continue
        statuses = [r[status_col] for r in tbl['rows'] if status_col < len(r)]
        total    = len(statuses)
        passed   = sum(1 for s in statuses if 'PASS' in s.upper())
        failed   = sum(1 for s in statuses if 'FAIL' in s.upper() and 'DEFERRED' not in s.upper())
        deferred = sum(1 for s in statuses if 'DEFERRED' in s.upper())
        rate     = f"{passed/total*100:.0f}%" if total else "N/A"
        summary_rows.append([label, total, passed, failed, deferred, rate])

    # Write summary rows
    for r_idx, row in enumerate(summary_rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = BODY_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border    = BORDER
            if c_idx == 3:   cell.fill = PASS_FILL
            elif c_idx == 4: cell.fill = FAIL_FILL
            elif c_idx == 5: cell.fill = DEFERRED_FILL

    # Totals row
    total_row = len(summary_rows) + 3
    totals = ["TOTAL",
              sum(r[1] for r in summary_rows),
              sum(r[2] for r in summary_rows),
              sum(r[3] for r in summary_rows),
              sum(r[4] for r in summary_rows),
              ""]
    if totals[1]:
        totals[5] = f"{totals[2]/totals[1]*100:.0f}%"
    for c_idx, val in enumerate(totals, start=1):
        cell = ws.cell(row=total_row, column=c_idx, value=val)
        cell.font   = Font(name="Calibri", bold=True, size=11)
        cell.fill   = PatternFill("solid", fgColor="D6E4F0")
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions['A'].width = 22
    for col in ['B','C','D','E','F']:
        ws.column_dimensions[col].width = 14

    # Bar chart
    if summary_rows:
        chart = BarChart()
        chart.type    = "col"
        chart.grouping = "clustered"
        chart.title   = "Test Results by Module"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Module"
        chart.width   = 22
        chart.height  = 14

        data_ref = Reference(ws,
                             min_col=3, max_col=5,
                             min_row=2, max_row=total_row - 1)
        cats_ref = Reference(ws, min_col=1,
                             min_row=3, max_row=total_row - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        # Colour the series
        series_colors = ["00B050", "FF0000", "808080"]
        for i, ser in enumerate(chart.series):
            ser.graphicalProperties.solidFill = series_colors[i]

        ws.add_chart(chart, f"A{total_row + 2}")


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    print(f"Reading {RTM_PATH} ...")
    text = RTM_PATH.read_text(encoding="utf-8")

    tables = parse_md_tables(text)
    print(f"Found {len(tables)} tables")

    wb = Workbook()
    # Remove default sheet placeholder (added after summary)
    default = wb.active
    wb.remove(default)

    # Build summary first (inserted at index 0 inside build_summary)
    build_summary(wb, tables)

    # Build one sheet per table
    for tbl in tables:
        sec = tbl['section']
        label = SECTION_MAP.get(sec, sec)
        # shorten sheet name to ≤31 chars (Excel limit); strip non-ASCII
        sheet_name = label.replace('—', '-').replace('–', '-')[:31]
        # deduplicate sheet names
        existing = [s.title for s in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:28] + f"_{existing.count(sheet_name)}"
        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, tbl, f"{sec} — Movie Review App RTM v2.1")
        print(f"  Sheet: {sheet_name} ({len(tbl['rows'])} rows)")

    wb.save(OUT_PATH)
    print(f"\n✅  Saved → {OUT_PATH}")


if __name__ == "__main__":
    main()
